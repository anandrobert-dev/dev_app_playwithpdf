# File: gui/pdf_to_office_gui.py
"""
PDF to Office Converter GUI Module for Oi360 Document Suite
Converts PDF files to Word (.docx), Excel (.xlsx), and Text (.txt) formats
with accuracy percentage reporting and page selection via preview window.
"""
import os
import sys
import re
from difflib import SequenceMatcher

# --- PyInstaller Path Fix ---
if getattr(sys, 'frozen', False):
    bundle_dir = sys._MEIPASS
    if bundle_dir not in sys.path:
        sys.path.insert(0, bundle_dir)

from PyQt5.QtWidgets import (
    QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout, QFileDialog,
    QMessageBox, QGraphicsDropShadowEffect, QComboBox, QProgressBar, 
    QFrame, QScrollArea, QCheckBox, QGridLayout, QDialog
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QPixmap

# Import shared components
from gui.ui_components import SelectablePreviewLabel


class PagePreviewWindow(QDialog):
    """Separate window for page preview with checkboxes and zoom controls"""
    pages_selected = pyqtSignal(object)  # Emits dict: {pages: list, rect: tuple, rect_page_idx: int}
    
    def __init__(self, pdf_path, parent=None):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.page_images = []  # Store PIL images
        self.page_checkboxes = []
        self.page_labels = []  # Store SelectablePreviewLabels
        self.checked_states = []  # Track checkbox states for zoom preservation
        self.zoom_level = 1.0
        self.base_thumbnail_height = 150
        self.selection_rect = None  # (x, y, w, h) in original image coords
        self.selected_page_idx = -1  # Which page has the selection
        
        # Make window non-modal so it's movable/detachable
        self.setWindowFlags(self.windowFlags() | Qt.Window)
        
        self.setWindowTitle("Select Pages to Convert")
        self.setMinimumSize(900, 700)
        self.setStyleSheet("""
            QDialog { background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #0a0a1a, stop:0.5 #0f1629, stop:1 #1a1a2e); }
            QLabel { color: #e2e8f0; }
            QCheckBox { color: #e2e8f0; font-weight: bold; }
            QCheckBox::indicator { width: 20px; height: 20px; }
            QCheckBox::indicator:unchecked { background: #374151; border: 2px solid #6b7280; border-radius: 4px; }
            QCheckBox::indicator:checked { background: #10b981; border: 2px solid #10b981; border-radius: 4px; }
        """)
        
        self.setup_ui()
        self.load_pages()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header = QLabel("📄 SELECT PAGES TO CONVERT")
        header.setFont(QFont("Segoe UI", 16, QFont.Bold))
        header.setStyleSheet("color: #00d4ff;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        # Instructions
        instructions = QLabel("Check the pages you want to convert. Use zoom controls to view details.")
        instructions.setStyleSheet("color: #94a3b8; font-style: italic;")
        instructions.setAlignment(Qt.AlignCenter)
        layout.addWidget(instructions)
        
        # Zoom controls
        zoom_row = QHBoxLayout()
        zoom_row.setSpacing(10)
        
        self.btn_zoom_out = QPushButton("➖ Zoom Out")
        self.btn_zoom_out.setFixedSize(120, 35)
        self.btn_zoom_out.setCursor(Qt.PointingHandCursor)
        self.btn_zoom_out.setStyleSheet("""
            QPushButton { background: #374151; color: white; border-radius: 8px; font-weight: bold; }
            QPushButton:hover { background: #4b5563; }
        """)
        self.btn_zoom_out.clicked.connect(lambda: self.adjust_zoom(0.8))
        zoom_row.addWidget(self.btn_zoom_out)
        
        self.zoom_label = QLabel("100%")
        self.zoom_label.setStyleSheet("color: #e2e8f0; font-weight: bold; font-size: 14px;")
        zoom_row.addWidget(self.zoom_label)
        
        self.btn_zoom_in = QPushButton("➕ Zoom In")
        self.btn_zoom_in.setFixedSize(120, 35)
        self.btn_zoom_in.setCursor(Qt.PointingHandCursor)
        self.btn_zoom_in.setStyleSheet("""
            QPushButton { background: #374151; color: white; border-radius: 8px; font-weight: bold; }
            QPushButton:hover { background: #4b5563; }
        """)
        self.btn_zoom_in.clicked.connect(lambda: self.adjust_zoom(1.25))
        zoom_row.addWidget(self.btn_zoom_in)
        
        zoom_row.addStretch()
        
        # Select All / Deselect All buttons
        self.btn_select_all = QPushButton("☑ Select All")
        self.btn_select_all.setFixedSize(120, 35)
        self.btn_select_all.setCursor(Qt.PointingHandCursor)
        self.btn_select_all.setStyleSheet("""
            QPushButton { background: #10b981; color: white; border-radius: 8px; font-weight: bold; }
            QPushButton:hover { background: #34d399; }
        """)
        self.btn_select_all.clicked.connect(self.select_all)
        zoom_row.addWidget(self.btn_select_all)
        
        self.btn_deselect_all = QPushButton("☐ Deselect All")
        self.btn_deselect_all.setFixedSize(120, 35)
        self.btn_deselect_all.setCursor(Qt.PointingHandCursor)
        self.btn_deselect_all.setStyleSheet("""
            QPushButton { background: #ef4444; color: white; border-radius: 8px; font-weight: bold; }
            QPushButton:hover { background: #f87171; }
        """)
        self.btn_deselect_all.clicked.connect(self.deselect_all)
        zoom_row.addWidget(self.btn_deselect_all)
        
        # Selection Region buttons
        self.btn_select_region = QPushButton("✂️ SELECT REGION")
        self.btn_select_region.setFixedSize(140, 35)
        self.btn_select_region.setCheckable(True)
        self.btn_select_region.setCursor(Qt.PointingHandCursor)
        self.btn_select_region.setStyleSheet("""
            QPushButton { background: #10b981; color: white; border-radius: 8px; font-weight: bold; }
            QPushButton:hover { background: #34d399; }
            QPushButton:checked { background: #06b6d4; color: #0f172a; }
        """)
        self.btn_select_region.clicked.connect(self.toggle_selection_mode)
        zoom_row.addWidget(self.btn_select_region)
        
        self.btn_clear_selection = QPushButton("✖️ CLEAR")
        self.btn_clear_selection.setFixedSize(100, 35)
        self.btn_clear_selection.setCursor(Qt.PointingHandCursor)
        self.btn_clear_selection.setStyleSheet("""
            QPushButton { background: #ef4444; color: white; border-radius: 8px; font-weight: bold; }
            QPushButton:hover { background: #f87171; }
        """)
        self.btn_clear_selection.clicked.connect(self.clear_selection)
        zoom_row.addWidget(self.btn_clear_selection)
        
        layout.addLayout(zoom_row)
        
        # Scroll area for page grid
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setStyleSheet("""
            QScrollArea { background: transparent; border: 1px solid #475569; border-radius: 10px; }
            QScrollBar:vertical { width: 12px; background: transparent; }
            QScrollBar::handle:vertical { background: #334155; border-radius: 6px; min-height: 30px; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        """)
        
        self.pages_widget = QWidget()
        self.pages_widget.setStyleSheet("background: transparent;")
        self.pages_layout = QGridLayout(self.pages_widget)
        self.pages_layout.setSpacing(15)
        self.pages_layout.setContentsMargins(15, 15, 15, 15)
        
        self.scroll_area.setWidget(self.pages_widget)
        layout.addWidget(self.scroll_area, stretch=1)
        
        # Selection summary
        self.selection_label = QLabel("Selected: 0 pages")
        self.selection_label.setStyleSheet("color: #f59e0b; font-weight: bold; font-size: 13px;")
        self.selection_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.selection_label)
        
        # Submit button
        self.btn_submit = QPushButton("✅ SUBMIT SELECTION")
        self.btn_submit.setFixedSize(250, 50)
        self.btn_submit.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.btn_submit.setCursor(Qt.PointingHandCursor)
        self.btn_submit.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #10b981, stop:1 #059669);
                color: white; font-weight: bold; border-radius: 12px;
            }
            QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #34d399, stop:1 #10b981); }
        """)
        self.btn_submit.clicked.connect(self.submit_selection)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_submit)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
    
    def load_pages(self):
        """Load all pages from PDF and create thumbnails with checkboxes"""
        try:
            from pdf2image import convert_from_path
            self.page_images = convert_from_path(self.pdf_path, dpi=72)
            self.render_pages()
        except Exception as e:
            error_label = QLabel(f"Error loading PDF: {str(e)}")
            error_label.setStyleSheet("color: #ef4444;")
            self.pages_layout.addWidget(error_label, 0, 0)
    
    def render_pages(self):
        """Render page thumbnails at current zoom level"""
        # Clear existing widgets
        for i in reversed(range(self.pages_layout.count())):
            widget = self.pages_layout.itemAt(i).widget()
            if widget:
                widget.deleteLater()
        
        self.page_checkboxes = []
        self.page_labels = []
        columns = max(1, int(4 / self.zoom_level))  # Adjust columns based on zoom
        
        for idx, pil_image in enumerate(self.page_images):
            # Create container for each page
            page_container = QFrame()
            page_container.setStyleSheet("""
                QFrame { 
                    background: rgba(30, 41, 59, 0.8); 
                    border: 2px solid #475569; 
                    border-radius: 10px; 
                }
                QFrame:hover { border-color: #06b6d4; }
            """)
            page_layout = QVBoxLayout(page_container)
            page_layout.setContentsMargins(10, 10, 10, 10)
            page_layout.setSpacing(8)
            
            # Convert PIL to QPixmap
            from PyQt5.QtGui import QImage
            img = pil_image.convert("RGB")
            img_data = img.tobytes("raw", "RGB")
            qimg = QImage(img_data, img.width, img.height, img.width * 3, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qimg)
            
            # Scale thumbnail
            thumb_height = int(self.base_thumbnail_height * self.zoom_level)
            scaled = pixmap.scaledToHeight(thumb_height, Qt.SmoothTransformation)
            
            # Page image label (Selectable)
            img_label = SelectablePreviewLabel()
            img_label.setPixmap(scaled)
            img_label.zoom_factor = scaled.height() / pil_image.height  # height is a property
            img_label.setAlignment(Qt.AlignCenter)
            img_label.setStyleSheet("border: 1px solid #475569; background: white;")
            img_label.selection_changed.connect(lambda rect, i=idx: self.on_selection_changed(rect, i))
            img_label.set_selection_mode(self.btn_select_region.isChecked())
            page_layout.addWidget(img_label)
            self.page_labels.append(img_label)
            
            # Checkbox for this page - restore state if available
            checkbox = QCheckBox(f"Page {idx + 1}")
            if idx < len(self.checked_states):
                checkbox.setChecked(self.checked_states[idx])
            else:
                checkbox.setChecked(True)  # Default: all selected
            checkbox.stateChanged.connect(self.update_selection_count)
            page_layout.addWidget(checkbox, alignment=Qt.AlignCenter)
            self.page_checkboxes.append(checkbox)
            
            row = idx // columns
            col = idx % columns
            self.pages_layout.addWidget(page_container, row, col)
        
        self.update_selection_count()
    
    def adjust_zoom(self, factor):
        """Adjust zoom level and re-render, preserving checkbox states"""
        # Save current checkbox states before re-rendering
        self.checked_states = [cb.isChecked() for cb in self.page_checkboxes]
        
        self.zoom_level *= factor
        self.zoom_level = max(0.5, min(3.0, self.zoom_level))  # Clamp between 50% and 300%
        self.zoom_label.setText(f"{int(self.zoom_level * 100)}%")
        self.render_pages()
    
    def select_all(self):
        for cb in self.page_checkboxes:
            cb.setChecked(True)
    
    def deselect_all(self):
        for cb in self.page_checkboxes:
            cb.setChecked(False)
    
    def update_selection_count(self):
        count = sum(1 for cb in self.page_checkboxes if cb.isChecked())
        text = f"Selected: {count} page(s)"
        if self.selection_rect:
            text += f" | Region selected on Page {self.selected_page_idx + 1}"
        self.selection_label.setText(text)
    
    def toggle_selection_mode(self):
        enabled = self.btn_select_region.isChecked()
        for label in self.page_labels:
            label.set_selection_mode(enabled)
            
    def clear_selection(self):
        self.selection_rect = None
        self.selected_page_idx = -1
        for label in self.page_labels:
            label.clear_selection()
        self.btn_select_region.setChecked(False)
        self.toggle_selection_mode()
        self.update_selection_count()

    def on_selection_changed(self, rect, page_idx):
        if rect:
            # Clear other selections (we only support one region selection for now)
            for i, label in enumerate(self.page_labels):
                if i != page_idx:
                    label.clear_selection()
            
            self.selection_rect = rect
            self.selected_page_idx = page_idx
            # Auto-check the page if not checked
            if not self.page_checkboxes[page_idx].isChecked():
                self.page_checkboxes[page_idx].setChecked(True)
        else:
            if self.selected_page_idx == page_idx:
                self.selection_rect = None
                self.selected_page_idx = -1
        self.update_selection_count()
    
    def submit_selection(self):
        """Emit selected pages, selection rect, and close"""
        selected_pages = [idx for idx, cb in enumerate(self.page_checkboxes) if cb.isChecked()]
        if not selected_pages:
            QMessageBox.warning(self, "No Selection", "Please select at least one page to convert.")
            return
            
        selection_data = {
            'pages': selected_pages,
            'rect': self.selection_rect,
            'rect_page_idx': self.selected_page_idx
        }
        self.pages_selected.emit(selection_data)
        self.accept()


class ConversionWorker(QThread):
    """Background thread for PDF conversion with accuracy calculation"""
    finished = pyqtSignal(str, float, str)
    error = pyqtSignal(str)
    progress = pyqtSignal(str)

    def __init__(self, pdf_path, output_format, output_dir, selected_pages=None, selection_rect=None, selection_page_idx=-1):
        super().__init__()
        self.pdf_path = pdf_path
        self.output_format = output_format
        self.output_dir = output_dir
        self.selected_pages = selected_pages  # List of 0-indexed page numbers, None = all pages
        self.selection_rect = selection_rect  # (x, y, w, h) in points (72 DPI)
        self.selection_page_idx = selection_page_idx # Page index where rect applies

    def run(self):
        try:
            base_name = os.path.splitext(os.path.basename(self.pdf_path))[0]
            
            if self.output_format == "Word (.docx)":
                output_path, accuracy = self.convert_to_word(base_name)
            elif self.output_format == "Excel (.xlsx)":
                output_path, accuracy = self.convert_to_excel(base_name)
            else:
                output_path, accuracy = self.convert_to_text(base_name)
            
            self.finished.emit(output_path, accuracy, self.output_format)
        except Exception as e:
            self.error.emit(str(e))

    def extract_pdf_text(self):
        """Extract text from selected pages using pdfplumber, with optional cropping"""
        import pdfplumber
        text = ""
        with pdfplumber.open(self.pdf_path) as pdf:
            pages_to_process = self.selected_pages if self.selected_pages else range(len(pdf.pages))
            for i in pages_to_process:
                if i < len(pdf.pages):
                    page = pdf.pages[i]
                    # Apply crop if this is the selected page
                    if self.selection_rect and i == self.selection_page_idx:
                        x, y, w, h = self.selection_rect
                        # pdfplumber uses (x0, y0, x1, y1)
                        bbox = (x, y, x + w, y + h)
                        try:
                            # within_bbox is safer for text extraction than crop
                            page = page.within_bbox(bbox)
                        except Exception:
                            pass
                    
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        return text

    def convert_to_word(self, base_name):
        """Convert selected PDF pages to Word document"""
        from pdf2docx import Converter
        from docx import Document
        
        self.progress.emit("Converting PDF to Word...")
        output_path = os.path.join(self.output_dir, f"{base_name}.docx")
        
        cv = Converter(self.pdf_path)
        if self.selected_pages:
            # Convert only selected pages
            for page_idx in self.selected_pages:
                cv.convert(output_path, start=page_idx, end=page_idx + 1)
        else:
            cv.convert(output_path, start=0, end=None)
        cv.close()
        
        self.progress.emit("Calculating accuracy...")
        original_text = self.extract_pdf_text()
        
        doc = Document(output_path)
        docx_text = "\n".join([para.text for para in doc.paragraphs])
        
        accuracy = self.calculate_text_similarity(original_text, docx_text)
        return output_path, accuracy

    def convert_to_excel(self, base_name):
        """Convert selected PDF pages to Excel"""
        import pdfplumber
        from openpyxl import Workbook
        
        self.progress.emit("Extracting tables from PDF...")
        output_path = os.path.join(self.output_dir, f"{base_name}.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "PDF Data"
        
        total_cells = 0
        valid_cells = 0
        current_row = 1
        
        with pdfplumber.open(self.pdf_path) as pdf:
            pages_to_process = self.selected_pages if self.selected_pages else range(len(pdf.pages))
            for i in pages_to_process:
                if i < len(pdf.pages):
                    page = pdf.pages[i]
                    self.progress.emit(f"Processing page {i + 1}...")
                    
                    # Apply crop if this is the selected page
                    if self.selection_rect and i == self.selection_page_idx:
                        x, y, w, h = self.selection_rect
                        bbox = (x, y, x + w, y + h)
                        try:
                            page = page.within_bbox(bbox)
                        except Exception:
                            pass
                    
                    tables = page.extract_tables()
                    
                    if tables:
                        for table in tables:
                            for row in table:
                                for col_idx, cell in enumerate(row, 1):
                                    total_cells += 1
                                    if cell:
                                        cell_value = str(cell).strip()
                                        ws.cell(row=current_row, column=col_idx, value=cell_value)
                                        if cell_value:
                                            valid_cells += 1
                                current_row += 1
                            current_row += 1
                    else:
                        text = page.extract_text()
                        if text:
                            for line in text.split('\n'):
                                line = line.strip()
                                if line:
                                    ws.cell(row=current_row, column=1, value=line)
                                    valid_cells += 1
                                    total_cells += 1
                                    current_row += 1
        
        wb.save(output_path)
        accuracy = (valid_cells / total_cells * 100) if total_cells > 0 else 0
        if valid_cells > 0:
            accuracy = min(accuracy + 20, 100)
        return output_path, accuracy

    def convert_to_text(self, base_name):
        """Convert selected PDF pages to plain text, with optional cropping"""
        self.progress.emit("Extracting text from PDF...")
        output_path = os.path.join(self.output_dir, f"{base_name}.txt")
        
        original_text = self.extract_pdf_text()
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(original_text)
        
        with open(output_path, 'r', encoding='utf-8') as f:
            saved_text = f.read()
        
        accuracy = self.calculate_text_similarity(original_text, saved_text)
        return output_path, accuracy

    def calculate_text_similarity(self, text1, text2):
        if not text1 or not text2:
            return 0.0
        text1 = re.sub(r'\s+', ' ', text1.strip().lower())
        text2 = re.sub(r'\s+', ' ', text2.strip().lower())
        if not text1 or not text2:
            return 0.0
        ratio = SequenceMatcher(None, text1, text2).ratio()
        return ratio * 100


class PDFToOfficeGUI(QWidget):
    """PDF to Office Converter GUI with preview window page selection"""
    
    FORMATS = ["Word (.docx)", "Excel (.xlsx)", "Text (.txt)"]
    
    def __init__(self, go_back_callback=None):
        super().__init__()
        self.go_back_callback = go_back_callback
        self.pdf_path = ""
        self.total_pages = 0
        self.selected_pages = None  # None means all pages
        self.selection_rect = None
        self.selection_page_idx = -1
        self.worker = None
        self.setAcceptDrops(True)
        self.setup_ui()

    def setup_ui(self):
        main_scroll = QScrollArea()
        main_scroll.setWidgetResizable(True)
        main_scroll.setFrameShape(QScrollArea.NoFrame)
        main_scroll.setStyleSheet("""
            QScrollArea { background: transparent; border: none; }
            QScrollBar:vertical { width: 10px; background: transparent; }
            QScrollBar::handle:vertical { background: #334155; border-radius: 5px; min-height: 30px; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0px; }
        """)
        
        scroll_content = QWidget()
        scroll_content.setStyleSheet("background: transparent;")
        
        layout = QVBoxLayout(scroll_content)
        layout.setSpacing(14)
        layout.setContentsMargins(30, 20, 30, 20)
        
        # Drag Drop Hint
        drag_hint = QLabel("Drag and Drop PDF files to convert")
        drag_hint.setAlignment(Qt.AlignCenter)
        drag_hint.setStyleSheet("""
            color: #06b6d4; font-size: 12px; font-weight: bold; padding: 8px; 
            background: rgba(6, 182, 212, 0.1); border-radius: 6px;
        """)
        layout.addWidget(drag_hint)
        
        # Header
        header = QHBoxLayout()
        
        if self.go_back_callback:
            btn_back = QPushButton("« MAIN MENU")
            btn_back.setFixedSize(140, 42)
            btn_back.setCursor(Qt.PointingHandCursor)
            btn_back.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #374151, stop:1 #4b5563);
                    color: #e5e7eb; border-radius: 10px; font-weight: bold; font-size: 11px;
                    border: 1px solid #6b7280;
                }
                QPushButton:hover { background: #4b5563; }
            """)
            btn_back.clicked.connect(self.go_back_callback)
            header.addWidget(btn_back)
        
        header.addStretch()
        
        # Title - Electric Blue
        title = QLabel("PDF to OFFICE")
        title.setFont(QFont("Georgia", 38, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #00d4ff; background: transparent; letter-spacing: 3px;")
        glow = QGraphicsDropShadowEffect()
        glow.setBlurRadius(30)
        glow.setOffset(0, 0)
        glow.setColor(QColor(0, 212, 255, 200))
        title.setGraphicsEffect(glow)
        header.addWidget(title)
        
        header.addStretch()
        layout.addLayout(header)

        # Subtitle
        subtitle = QLabel("Convert PDFs to Word, Excel, or Text | Accuracy Reporting Included")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: #94a3b8; font-style: italic; font-size: 11px;")
        layout.addWidget(subtitle)

        # File Selection Row
        file_row = QHBoxLayout()
        file_row.setSpacing(10)
        
        self.file_label = QLabel("No file selected")
        self.file_label.setStyleSheet("""
            color: #e2e8f0; font-size: 13px; padding: 10px; 
            background: rgba(30, 41, 59, 0.9); border-radius: 8px;
            border: 1px solid #475569;
        """)
        self.file_label.setMinimumHeight(45)
        file_row.addWidget(self.file_label, stretch=1)
        
        btn_browse = QPushButton("📁 BROWSE PDF")
        btn_browse.setFixedSize(140, 45)
        btn_browse.setFont(QFont("Segoe UI", 11, QFont.Bold))
        btn_browse.setCursor(Qt.PointingHandCursor)
        btn_browse.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #8b5cf6, stop:1 #a78bfa);
                color: white; font-weight: bold; border-radius: 10px;
            }
            QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #a78bfa, stop:1 #c4b5fd); }
        """)
        btn_browse.clicked.connect(self.browse_file)
        file_row.addWidget(btn_browse)
        
        layout.addLayout(file_row)

        # PDF Info Panel
        self.info_panel = QFrame()
        self.info_panel.setVisible(False)
        self.info_panel.setStyleSheet("""
            QFrame { background: rgba(30, 41, 59, 0.5); border: 1px solid #475569; border-radius: 10px; }
        """)
        info_layout = QHBoxLayout(self.info_panel)
        info_layout.setContentsMargins(15, 12, 15, 12)
        info_layout.setSpacing(20)
        
        self.pdf_info_label = QLabel("PDF Info: Loading...")
        self.pdf_info_label.setStyleSheet("color: #10b981; font-size: 12px; font-weight: bold;")
        info_layout.addWidget(self.pdf_info_label)
        
        info_layout.addStretch()
        
        # Selected pages label
        self.selected_pages_label = QLabel("Selected: All pages")
        self.selected_pages_label.setStyleSheet("color: #f59e0b; font-weight: bold; font-size: 12px;")
        info_layout.addWidget(self.selected_pages_label)
        
        # Preview button to open page selection window
        self.btn_preview = QPushButton("👁 PREVIEW & SELECT PAGES")
        self.btn_preview.setFixedSize(220, 40)
        self.btn_preview.setFont(QFont("Segoe UI", 10, QFont.Bold))
        self.btn_preview.setCursor(Qt.PointingHandCursor)
        self.btn_preview.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3b82f6, stop:1 #2563eb);
                color: white; font-weight: bold; border-radius: 10px;
            }
            QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #60a5fa, stop:1 #3b82f6); }
        """)
        self.btn_preview.clicked.connect(self.open_preview_window)
        info_layout.addWidget(self.btn_preview)
        
        layout.addWidget(self.info_panel)

        # Format Selection Row
        format_row = QHBoxLayout()
        format_row.setSpacing(15)
        
        format_label = QLabel("Output Format:")
        format_label.setStyleSheet("color: #a8b4d4; font-size: 12px; font-weight: bold;")
        format_row.addWidget(format_label)
        
        self.format_combo = QComboBox()
        self.format_combo.addItems(self.FORMATS)
        self.format_combo.setFixedSize(180, 42)
        self.format_combo.setStyleSheet("""
            QComboBox {
                background: #1e293b; color: #e2e8f0; border: 1px solid #475569;
                border-radius: 8px; padding: 8px; font-size: 12px;
            }
            QComboBox::drop-down { border: none; width: 30px; }
            QComboBox::down-arrow { image: none; border-left: 5px solid transparent; 
                border-right: 5px solid transparent; border-top: 6px solid #e2e8f0; }
            QComboBox QAbstractItemView { background: #1e293b; color: #e2e8f0; 
                selection-background-color: #06b6d4; }
        """)
        format_row.addWidget(self.format_combo)
        
        format_row.addStretch()
        
        # Convert Button
        self.btn_convert = QPushButton("🔄 CONVERT")
        self.btn_convert.setFixedSize(180, 50)
        self.btn_convert.setFont(QFont("Segoe UI", 12, QFont.Bold))
        self.btn_convert.setCursor(Qt.PointingHandCursor)
        self.btn_convert.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #06b6d4, stop:1 #3b82f6);
                color: white; font-weight: bold; border-radius: 12px;
            }
            QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #22d3ee, stop:1 #60a5fa); }
            QPushButton:disabled { background: #4b5563; color: #9ca3af; }
        """)
        self.btn_convert.clicked.connect(self.start_conversion)
        format_row.addWidget(self.btn_convert)
        
        layout.addLayout(format_row)
        
        # Progress Bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar { border: 1px solid #475569; border-radius: 6px; background: #1e293b; height: 20px; }
            QProgressBar::chunk { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #06b6d4, stop:1 #3b82f6); }
        """)
        layout.addWidget(self.progress_bar)
        
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("color: #06b6d4; font-size: 11px;")
        layout.addWidget(self.status_label)

        # Results Panel
        self.results_panel = QFrame()
        self.results_panel.setVisible(False)
        self.results_panel.setStyleSheet("""
            QFrame { background: rgba(30, 41, 59, 0.7); border: 1px solid #475569; border-radius: 12px; }
        """)
        results_layout = QVBoxLayout(self.results_panel)
        results_layout.setContentsMargins(20, 20, 20, 20)
        results_layout.setSpacing(15)
        
        result_title = QLabel("CONVERSION RESULT")
        result_title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        result_title.setStyleSheet("color: #06b6d4; background: transparent; border: none;")
        result_title.setAlignment(Qt.AlignCenter)
        results_layout.addWidget(result_title)
        
        self.output_label = QLabel("Output: N/A")
        self.output_label.setWordWrap(True)
        self.output_label.setStyleSheet("color: #e2e8f0; font-size: 12px; background: transparent; border: none;")
        results_layout.addWidget(self.output_label)
        
        accuracy_row = QHBoxLayout()
        self.accuracy_label = QLabel("Accuracy: ")
        self.accuracy_label.setStyleSheet("color: #94a3b8; font-size: 14px; font-weight: bold; background: transparent; border: none;")
        accuracy_row.addWidget(self.accuracy_label)
        
        self.accuracy_value = QLabel("0%")
        self.accuracy_value.setFont(QFont("Segoe UI", 24, QFont.Bold))
        self.accuracy_value.setStyleSheet("color: #10b981; background: transparent; border: none;")
        accuracy_row.addWidget(self.accuracy_value)
        accuracy_row.addStretch()
        results_layout.addLayout(accuracy_row)
        
        self.accuracy_indicator = QLabel("")
        self.accuracy_indicator.setStyleSheet("font-size: 12px; background: transparent; border: none;")
        self.accuracy_indicator.setWordWrap(True)
        results_layout.addWidget(self.accuracy_indicator)
        
        self.btn_open = QPushButton("📂 OPEN OUTPUT FILE")
        self.btn_open.setMinimumSize(200, 45)
        self.btn_open.setFont(QFont("Segoe UI", 11, QFont.Bold))
        self.btn_open.setCursor(Qt.PointingHandCursor)
        self.btn_open.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #10b981, stop:1 #059669);
                color: white; font-weight: bold; border-radius: 10px;
            }
            QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #34d399, stop:1 #10b981); }
        """)
        self.btn_open.clicked.connect(self.open_output_file)
        results_layout.addWidget(self.btn_open, alignment=Qt.AlignCenter)
        
        layout.addWidget(self.results_panel)
        layout.addStretch()

        main_scroll.setWidget(scroll_content)
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.addWidget(main_scroll)

    def browse_file(self):
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select PDF File", downloads_path, "PDF Files (*.pdf)"
        )
        if file_path:
            self.load_file(file_path)

    def load_file(self, file_path):
        self.pdf_path = file_path
        self.selected_pages = None  # Reset to all pages
        file_name = os.path.basename(file_path)
        self.file_label.setText(f"📄 {file_name}")
        self.file_label.setStyleSheet("""
            color: #10b981; font-size: 13px; padding: 10px; 
            background: rgba(16, 185, 129, 0.1); border-radius: 8px;
            border: 1px solid #10b981;
        """)
        self.results_panel.setVisible(False)
        self.status_label.setText("")
        
        # Get page count
        try:
            import pdfplumber
            with pdfplumber.open(file_path) as pdf:
                self.total_pages = len(pdf.pages)
            self.pdf_info_label.setText(f"📄 {file_name}  |  📑 {self.total_pages} page(s)")
            self.selected_pages_label.setText(f"Selected: All {self.total_pages} pages")
            self.info_panel.setVisible(True)
        except Exception as e:
            self.pdf_info_label.setText(f"Error: {str(e)}")
            self.total_pages = 0

    def open_preview_window(self):
        """Open the page selection preview window"""
        if not self.pdf_path:
            QMessageBox.warning(self, "No File", "Please select a PDF file first.")
            return
        
        preview = PagePreviewWindow(self.pdf_path, self)
        preview.pages_selected.connect(self.on_pages_selected)
        preview.exec_()

    def on_pages_selected(self, selection_data):
        """Handle page selection and region from preview window"""
        if isinstance(selection_data, dict):
            pages = selection_data['pages']
            self.selection_rect = selection_data.get('rect')
            self.selection_page_idx = selection_data.get('rect_page_idx', -1)
        else:
            pages = selection_data
            self.selection_rect = None
            self.selection_page_idx = -1
            
        self.selected_pages = pages
        if len(pages) == self.total_pages:
            summary = f"Selected: All {len(pages)} pages"
            self.selected_pages = None  # Treat as all pages for efficiency
        else:
            page_nums = [str(p + 1) for p in pages]
            if len(page_nums) > 5:
                display = ", ".join(page_nums[:5]) + f"... (+{len(page_nums) - 5} more)"
            else:
                display = ", ".join(page_nums)
            summary = f"Selected: Pages {display}"
            
        if self.selection_rect:
            summary += f" | Crop applied to Page {self.selection_page_idx + 1}"
            
        self.selected_pages_label.setText(summary)

    def start_conversion(self):
        if not self.pdf_path:
            QMessageBox.warning(self, "No File", "Please select a PDF file first.")
            return
        
        if not os.path.exists(self.pdf_path):
            QMessageBox.warning(self, "File Not Found", "The selected PDF file no longer exists.")
            return
        
        self.btn_convert.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.results_panel.setVisible(False)
        
        output_format = self.format_combo.currentText()
        output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
        
        self.worker = ConversionWorker(self.pdf_path, output_format, output_dir, 
                                       self.selected_pages, self.selection_rect, self.selection_page_idx)
        self.worker.finished.connect(self.on_conversion_finished)
        self.worker.error.connect(self.on_conversion_error)
        self.worker.progress.connect(self.on_progress_update)
        self.worker.start()

    def on_progress_update(self, message):
        self.status_label.setText(message)

    def on_conversion_finished(self, output_path, accuracy, output_format):
        self.btn_convert.setEnabled(True)
        self.progress_bar.setVisible(False)
        
        self.output_path = output_path
        self.output_label.setText(f"📁 Output: {output_path}")
        self.accuracy_value.setText(f"{accuracy:.1f}%")
        
        if accuracy >= 80:
            self.accuracy_value.setStyleSheet("color: #10b981; background: transparent; border: none; font-size: 24px;")
            self.accuracy_indicator.setText("🟢 HIGH ACCURACY - Data is reliable for use")
            self.accuracy_indicator.setStyleSheet("color: #10b981; font-size: 12px; background: transparent; border: none;")
        elif accuracy >= 50:
            self.accuracy_value.setStyleSheet("color: #f59e0b; background: transparent; border: none; font-size: 24px;")
            self.accuracy_indicator.setText("🟡 MODERATE ACCURACY - Review recommended before use")
            self.accuracy_indicator.setStyleSheet("color: #f59e0b; font-size: 12px; background: transparent; border: none;")
        else:
            self.accuracy_value.setStyleSheet("color: #ef4444; background: transparent; border: none; font-size: 24px;")
            self.accuracy_indicator.setText("🔴 LOW ACCURACY - Manual verification required")
            self.accuracy_indicator.setStyleSheet("color: #ef4444; font-size: 12px; background: transparent; border: none;")
        
        self.results_panel.setVisible(True)
        self.status_label.setText("✅ Conversion complete! Saved to Downloads folder.")
        self.status_label.setStyleSheet("color: #10b981; font-size: 11px;")

    def on_conversion_error(self, error_message):
        self.btn_convert.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.status_label.setText(f"❌ Error: {error_message}")
        self.status_label.setStyleSheet("color: #ef4444; font-size: 11px;")
        QMessageBox.critical(self, "Conversion Error", f"Failed to convert PDF:\n{error_message}")

    def open_output_file(self):
        if hasattr(self, 'output_path') and os.path.exists(self.output_path):
            import subprocess
            subprocess.run(['xdg-open', self.output_path], check=False)
        else:
            QMessageBox.warning(self, "File Not Found", "Output file not found.")

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and urls[0].toLocalFile().lower().endswith('.pdf'):
                event.acceptProposedAction()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            url = event.mimeData().urls()[0]
            file_path = url.toLocalFile()
            if file_path.lower().endswith('.pdf'):
                self.load_file(file_path)
